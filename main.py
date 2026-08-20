#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Parse the 账务说明 (account summary) and 交易明细 (transactions) sections
of a 中国农业银行 credit-card e-statement .eml, then reconcile the two.

The transaction details are summed per category (还款/费用/消费/退货) and
compared against the 账务说明 amounts:
  - 本期账单金额            == 消费 + 费用
  - 本期还款、退货金额       == 还款 + 退货
  - 本期应还 - 本期溢缴款    == 上期应还 - 上期溢缴 + 本期账单 - 本期还款退货 - 本期调整
If any check fails, a ValueError is raised (script exits non-zero).

After the check passes, the transactions are grouped by 卡号后四位 and each
card gets its own sheet plus a per-card total (待还款总额 = 消费+费用 − 退货;
还款 belongs to the previous billing cycle and is not netted in). The per-card
totals are sanity-checked against the whole-account totals.

Outputs multiple .xlsx files, one per table/section:
  {prefix}_summary.xlsx        账务说明
  {prefix}_transactions.xlsx   全部交易明细
  {prefix}_cards.xlsx          按卡汇总
  {prefix}_card_{卡号}.xlsx     每个卡的交易明细

Usage:
  uv run main.py 2608                # -> reads bills/2608.eml, writes xlsx into bills/2608/
  uv run main.py path/to/2608.eml    # -> writes xlsx next to the .eml
"""
import argparse
import base64
import email
import io
import os
import re
import sys
from collections import Counter
from email.header import decode_header, make_header
from html.parser import HTMLParser

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from wcwidth import wcswidth

# =============================================================================
# 常量
# =============================================================================

CATEGORIES = ('还款', '费用', '消费', '退货', '调整')
DEBIT_TYPES = ('消费', '费用')   # 本期新增支出，计入待还款
REFUND_TYPE = '退货'             # 冲减本期支出
PAYMENT_TYPE = '还款'            # 偿还上一账单周期，不计入待还款

DATA_DIR = 'bills'               # 原始 .eml 所在目录；输出到 <DATA_DIR>/<期号>/
EPS = 0.005                      # 金额比较容差（分）

TX_BASE_HEADER = ['交易日期', '入账日期', '卡号后四位', '交易描述', '类型']
CARD_HEADER = ['卡号后四位', '交易笔数', '消费(含费用)', '退货', '还款(上期)', '待还款总额', '还款日']
SUMMARY_KEYS = ['币种', '本期应还金额', '本期账户溢缴款', '上期账单应还金额',
                '上期账户溢缴款', '本期账单金额', '本期还款、退货金额', '本期调整金额']


# =============================================================================
# HTML 表格抽取
# =============================================================================

class TableExtractor(HTMLParser):
    """把每个 <table> 收集成行列结构：tables = [[row, ...], ...]，row = [cell, ...]."""

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.tables = []
        self.stack = []  # 每个元素: {'rows': [], 'row': None, 'cell': [], 'incell': False}

    @classmethod
    def from_html(cls, html):
        parser = cls()
        parser.feed(html)
        return parser

    def handle_starttag(self, tag, attrs):
        if tag == 'table':
            self.stack.append({'rows': [], 'row': None, 'cell': [], 'incell': False})
        elif self.stack:
            ctx = self.stack[-1]
            if tag == 'tr':
                ctx['row'] = []
            elif tag in ('td', 'th'):
                ctx['cell'] = []
                ctx['incell'] = True

    def handle_data(self, data):
        if self.stack and self.stack[-1]['incell']:
            self.stack[-1]['cell'].append(data)

    def handle_endtag(self, tag):
        if not self.stack:
            return
        ctx = self.stack[-1]
        if tag in ('td', 'th'):
            if ctx['incell']:
                ctx['row'].append(''.join(ctx['cell']).strip())
                ctx['incell'] = False
        elif tag == 'tr':
            if ctx['row'] is not None:
                if any(c.strip() for c in ctx['row']):
                    ctx['rows'].append(ctx['row'])
                ctx['row'] = None
        elif tag == 'table':
            self.tables.append(self.stack.pop()['rows'])


def norm(s):
    """单元格文本规范化：NBSP/换行 -> 空格，合并空白。"""
    s = s.replace('\xa0', ' ')
    s = s.replace('\n', ' ')
    return re.sub(r'\s+', ' ', s).strip()


def cell(row, i):
    return norm(row[i]) if i < len(row) else ''


# =============================================================================
# 邮件解析
# =============================================================================

def parse_eml(path):
    """从文件读取 .eml，返回 (email_info, html_body)。"""
    return parse_eml_bytes(open(path, 'rb').read())


def parse_eml_bytes(raw):
    """解析 .eml 字节流，返回 (email_info, html_body)。"""
    msg = email.message_from_bytes(raw)

    def dec(field):
        try:
            return str(make_header(decode_header(msg.get(field, ''))))
        except Exception:
            return msg.get(field, '')

    info = {'主题': dec('Subject'), '发件人': dec('From'),
            '收件人': dec('To'), '发送时间': dec('Date')}

    body = ''
    for part in msg.walk():
        if part.get_content_type() == 'text/html':
            body = part.get_payload(decode=True).decode('utf-8', 'ignore')
            break

    if not body:  # 兜底：手动切分头部/正文并 base64 解码
        head, sep, b = raw.partition(b'\n\n')
        if not sep:
            head, sep, b = raw.partition(b'\r\n\r\n')
        body = base64.b64decode(re.sub(rb'\s+', b'', b)).decode('utf-8', 'ignore')

    return info, body


# =============================================================================
# 业务提取：账户信息 / 账务说明 / 交易明细
# =============================================================================

def _category_from_label_table(tbl):
    """交易分组标签表（单行、含分类词、无 6 位日期）-> 分类名，否则 None。"""
    if len(tbl) != 1:
        return None
    labels = [norm(c) for c in tbl[0]]
    cats = [c for c in labels if c in CATEGORIES]
    if cats and not any(re.fullmatch(r'\d{6}', c) for c in labels):
        return cats[-1]
    return None


def _due_date_from(row):
    """账户信息里的到期还款日，否则 None。"""
    if len(row) >= 2 and '到期还款日' in cell(row, 0) and 'Payment Due Date' in cell(row, 0):
        return cell(row, 1)
    return None


def _summary_from(row):
    """账务说明数据行（8 格、首格人民币）-> {key: value}，否则 None。"""
    if len(row) == len(SUMMARY_KEYS) and norm(row[0]).startswith('人民币'):
        return dict(zip(SUMMARY_KEYS, map(norm, row)))
    return None


def _transaction_from(row, category):
    """交易明细行（6+ 格、首格 6 位日期）-> 交易 dict，否则 None。"""
    if len(row) >= 6 and re.fullmatch(r'\d{6}', norm(row[0])):
        return {'交易日期': norm(row[0]), '入账日期': norm(row[1]),
                '卡号后四位': norm(row[2]), '交易描述': norm(row[3]),
                '交易金额/币种': norm(row[4]), '入账金额/币种': norm(row[5]),
                '类型': category}
    return None


def extract_summary_and_txns(tables):
    """解析账户信息(到期还款日)、账务说明和交易明细。"""
    account, summary, txns = {}, {}, []
    category = ''

    for tbl in tables:
        if (c := _category_from_label_table(tbl)) is not None:
            category = c
            continue
        for row in tbl:
            if (v := _due_date_from(row)) is not None:
                account['到期还款日'] = v
            elif (s := _summary_from(row)) is not None:
                summary.update(s)
            elif (t := _transaction_from(row, category)) is not None:
                txns.append(t)

    return account, summary, txns


# =============================================================================
# 金额处理
# =============================================================================

def to_amount(s):
    """解析金额字符串 '1,234.56' / '-50.00' / '343.63/CNY' -> float。"""
    s = re.sub(r'[^\d.-]', '', s or '')
    return float(s) if s else 0.0


def split_amount(s):
    """'343.63/CNY' -> (343.63, 'CNY')；无法解析返回 (None, '')。"""
    m = re.match(r'^\s*([-\d.,]+)\s*/\s*([A-Za-z]{2,4})\s*$', s or '')
    if not m:
        return None, ''
    return float(m.group(1).replace(',', '')), m.group(2).upper()


def transaction_columns(txns):
    """交易表：金额按币种拆成独立列，如 交易金额/CNY、入账金额/USD……值为纯数字。"""
    curs = sorted({c for t in txns
                   for c in (split_amount(t['交易金额/币种'])[1], split_amount(t['入账金额/币种'])[1])
                   if c})
    header = list(TX_BASE_HEADER) + [f'交易金额/{c}' for c in curs] + [f'入账金额/{c}' for c in curs]
    rows = []
    for t in txns:
        ta, tc = split_amount(t['交易金额/币种'])
        sa, sc = split_amount(t['入账金额/币种'])
        row = [t[k] for k in TX_BASE_HEADER]
        row += [ta if c == tc else None for c in curs]
        row += [sa if c == sc else None for c in curs]
        rows.append(row)
    return header, rows


# =============================================================================
# 核对：交易统计 vs 账务说明
# =============================================================================

def verify(summary, txns):
    """把交易按类型汇总，与账务说明比对；不一致抛出 ValueError。"""
    errors = []

    def check(label, actual, expected):
        if abs(actual - expected) > EPS:
            errors.append(f'{label}: 交易统计 {actual:,.2f} ≠ 账务说明 {expected:,.2f}')

    sums, counts = {}, Counter()
    for t in txns:
        sums[t['类型']] = sums.get(t['类型'], 0.0) + abs(to_amount(t['入账金额/币种']))
        counts[t['类型']] += 1

    if counts.get('', 0):
        errors.append(f'存在 {counts[""]} 笔未能归类的交易，金额合计 {sums.get("", 0.0):,.2f}')

    new_charge = sums.get('消费', 0.0) + sums.get('费用', 0.0)
    payments = sums.get('还款', 0.0) + sums.get('退货', 0.0)

    check('本期账单金额(消费+费用)', new_charge, to_amount(summary.get('本期账单金额')))
    check('本期还款、退货金额(还款+退货)', payments, to_amount(summary.get('本期还款、退货金额')))

    # 主余额等式（对账单表头）：
    # 应还 - 溢缴 = 上期应还 - 上期溢缴 + 本期账单 - 本期还款退货 - 本期调整
    nb = to_amount(summary.get('本期应还金额'))
    dep = to_amount(summary.get('本期账户溢缴款'))
    prev = to_amount(summary.get('上期账单应还金额'))
    prev_dep = to_amount(summary.get('上期账户溢缴款'))
    adj = to_amount(summary.get('本期调整金额'))
    lhs = nb - dep
    rhs = prev - prev_dep + new_charge - payments - adj
    if abs(lhs - rhs) > EPS:
        errors.append(
            f'余额等式不成立: 本期应还-溢缴 = {lhs:,.2f}, 应为 '
            f'{prev:,.2f}(上期应还) - {prev_dep:,.2f}(上期溢缴) + {new_charge:,.2f}(本期账单) '
            f'- {payments:,.2f}(本期还款退货) - {adj:,.2f}(本期调整) = {rhs:,.2f}')

    if errors:
        raise ValueError('账务说明与交易明细核对不一致:\n  ' + '\n'.join(errors))

    return sums, counts


# =============================================================================
# 按卡汇总
# =============================================================================

def group_by_card(txns):
    """按卡号后四位分组交易。"""
    cards = {}
    for t in txns:
        c = t['卡号后四位'] or '(未知)'
        cards.setdefault(c, []).append(t)
    return cards


def card_stats(card_txns):
    """单卡统计。待还款总额(本期) = 消费+费用 − 退货。"""
    debit = refund = payment = 0.0
    for t in card_txns:
        amt = abs(to_amount(t['入账金额/币种']))
        if t['类型'] in DEBIT_TYPES:
            debit += amt
        elif t['类型'] == REFUND_TYPE:
            refund += amt
        elif t['类型'] == PAYMENT_TYPE:
            payment += amt
    return debit, refund, payment, debit - refund


def build_card_summary(cards, due_date):
    """生成按卡汇总行（含合计行）和每卡汇总 dict。返回 (rows, by_card, totals)。"""
    rows, by_card = [], {}
    totals = {'debit': 0.0, 'refund': 0.0, 'payment': 0.0}

    for c in sorted(cards):
        debit, refund, payment, net = card_stats(cards[c])
        totals['debit'] += debit
        totals['refund'] += refund
        totals['payment'] += payment
        row = [c, len(cards[c]), f'{debit:,.2f}', f'{refund:,.2f}',
               f'{payment:,.2f}', f'{net:,.2f}', due_date]
        rows.append(row)
        by_card[c] = row

    totals['net'] = totals['debit'] - totals['refund']
    rows.append(['合计', sum(len(v) for v in cards.values()),
                 f"{totals['debit']:,.2f}", f"{totals['refund']:,.2f}",
                 f"{totals['payment']:,.2f}", f"{totals['net']:,.2f}", ''])
    return rows, by_card, totals


def verify_card_totals(sums, totals):
    """按卡分组合计必须等于全账户合计，否则数据有误。"""
    if abs(totals['debit'] - (sums.get('消费', 0.0) + sums.get('费用', 0.0))) > EPS or \
       abs(totals['refund'] - sums.get('退货', 0.0)) > EPS or \
       abs(totals['payment'] - sums.get('还款', 0.0)) > EPS:
        raise ValueError('按卡分组的消费/退货/还款合计与全账户合计不一致。')


# =============================================================================
# xlsx 渲染
# =============================================================================

def build_summary_rows(info, summary):
    rows = [['项目', '内容']]
    for k in ('主题', '发件人', '收件人', '发送时间'):
        if info.get(k):
            rows.append([k, info[k]])
    for k in SUMMARY_KEYS:
        if summary.get(k):
            rows.append([k, summary[k]])
    return rows


def fill_sheet(ws, header, rows, start_row=1):
    """把表头+数据写进 sheet，带基础格式。

    表头以 交易金额/ 或 入账金额/ 开头的列按数字写入，并套用千分位两位小数格式。
    """
    numeric = {j for j, h in enumerate(header, 1) if h.startswith(('交易金额/', '入账金额/'))}
    for j, h in enumerate(header, 1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='D9D9D9')
    for i, r in enumerate(rows, start_row + 1):
        for j, v in enumerate(r, 1):
            c = ws.cell(row=i, column=j, value=v)
            if j in numeric and isinstance(v, (int, float)):
                c.number_format = '#,##0.00'
    ws.freeze_panes = f'A{start_row + 1}'
    for j, h in enumerate(header, 1):
        width = max((len(str(r[j - 1])) if j <= len(r) else 0) for r in rows) if rows else 0
        ws.column_dimensions[get_column_letter(j)].width = min(max(max(len(str(h)), width) + 2, 8), 60)


def xlsx_bytes(title, header, rows, head_block=None):
    """生成单 sheet 的 xlsx 字节流（供网页端下载 / 内存使用）。

    head_block = (head_header, head_rows)：写在文件顶部，空 2 行后再写主表，
    供每个卡的明细文件显示该卡自己的汇总。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title
    if head_block:
        head_header, head_rows = head_block
        fill_sheet(ws, head_header, head_rows)
        fill_sheet(ws, header, rows, start_row=len(head_rows) + 4)
    else:
        fill_sheet(ws, header, rows)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def write_xlsx(path, title, header, rows, head_block=None):
    """写一个单 sheet 的 xlsx 文件。"""
    with open(path, 'wb') as f:
        f.write(xlsx_bytes(title, header, rows, head_block))


# =============================================================================
# CLI / 编排
# =============================================================================

def out_name(prefix, suffix):
    """拼接输出文件名；prefix 为空则直接用 suffix。"""
    return f'{prefix}_{suffix}' if prefix else suffix


def resolve_input(arg, outdir=None):
    """把位置参数解析为 (eml_path, prefix, outdir)。

    两种形式：
      - 期号 "2608"        -> 读 bills/2608.eml，输出到 bills/2608/
      - .eml 完整路径      -> 输出到其所在目录
    """
    if arg.lower().endswith('.eml'):
        eml_path = arg
        prefix = os.path.splitext(os.path.basename(arg))[0]
        outdir = outdir or os.path.dirname(os.path.abspath(arg))
    else:
        name = os.path.basename(arg.rstrip('/'))
        eml_path = os.path.join(DATA_DIR, name + '.eml')
        prefix = name
        outdir = outdir or os.path.join(DATA_DIR, name)
    if not os.path.isfile(eml_path):
        raise SystemExit(f'找不到账单文件: {eml_path}')
    return eml_path, prefix, outdir


def write_outputs(outdir, prefix, info, summary, txns, cards, card_rows, card_summary):
    """写所有 xlsx，返回 (paths, card_paths)。"""
    txn_header, txn_rows = transaction_columns(txns)
    paths = {
        '摘要': os.path.join(outdir, out_name(prefix, 'summary.xlsx')),
        '交易': os.path.join(outdir, out_name(prefix, 'transactions.xlsx')),
        '按卡汇总': os.path.join(outdir, out_name(prefix, 'cards.xlsx')),
    }
    write_xlsx(paths['摘要'], '账务说明', ['项目', '内容'], build_summary_rows(info, summary)[1:])
    write_xlsx(paths['交易'], '交易明细', txn_header, txn_rows)
    write_xlsx(paths['按卡汇总'], '按卡汇总', CARD_HEADER, card_rows)

    card_paths = {}
    for c in sorted(cards):
        p = os.path.join(outdir, out_name(prefix, f'card_{c}.xlsx'))
        card_paths[c] = p
        h, r = transaction_columns(cards[c])
        write_xlsx(p, '交易明细', h, r, head_block=(CARD_HEADER, [card_summary[c]]))
    return paths, card_paths


def _disp_width(s):
    """终端显示宽度（CJK 全角按 2 格），交给 wcwidth 计算。"""
    w = wcswidth(s)
    return w if w >= 0 else len(s)


def render_table(header, rows):
    """按显示宽度用空格补齐，返回对齐好的表格行列表（中文按 2 格计）。"""
    all_rows = [header] + [list(map(str, r)) for r in rows]
    n = len(header)
    widths = [0] * n
    for row in all_rows:
        for c in range(n):
            v = row[c] if c < len(row) else ''
            widths[c] = max(widths[c], _disp_width(v))
    out = []
    for row in all_rows:
        cells = []
        for c in range(n):
            v = row[c] if c < len(row) else ''
            cells.append(v + ' ' * max(0, widths[c] - _disp_width(v)))
        out.append('  '.join(cells).rstrip())
    return out


def build_report_text(paths, card_paths, txns, counts, sums, card_rows, totals, summary):
    """把核对结果和按卡汇总拼成多行文本；paths 为 None 时省略文件路径。"""
    lines = []
    if paths:
        lines.append(f'摘要: {paths["摘要"]}')
        lines.append(f'交易: {paths["交易"]}')
        lines.append(f'按卡汇总: {paths["按卡汇总"]}')
        for c in sorted(card_paths):
            lines.append(f'明细: {card_paths[c]}')
    lines.append(f'交易笔数: {len(txns)}  ' + '  '.join(f'{k}={v}' for k, v in counts.items()))
    lines.append('按类型统计(入账金额绝对值):')
    for k in CATEGORIES:
        if k in counts:
            lines.append(f'  {k:<4} {counts[k]:>3} 笔  合计 {sums.get(k, 0.0):>12,.2f}')
    lines.append('按卡汇总[待还款总额(本期) = 消费+费用 − 退货; 还款为上期还款, 不计入]:')
    lines += render_table(CARD_HEADER, card_rows)
    lines.append(f'注: 各卡待还款合计 {totals["net"]:,.2f} − 本期调整 '
                 f'{to_amount(summary.get("本期调整金额")):,.2f} = '
                 f'本期应还金额 {to_amount(summary.get("本期应还金额")):,.2f}')
    lines.append('核对通过: 交易明细与账务说明一致。')
    return '\n'.join(lines)


def print_report(paths, card_paths, txns, counts, sums, card_rows, totals, summary):
    print(build_report_text(paths, card_paths, txns, counts, sums, card_rows, totals, summary))


def parse_args(argv):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('eml', help='billing code like "2608" (=> bills/2608.eml) or a path to a .eml')
    ap.add_argument('--outdir', default=None, help='output directory (default: next to input)')
    ap.add_argument('--prefix', default=None, help='output filename prefix')
    return ap.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    eml_path, prefix, outdir = resolve_input(args.eml, args.outdir)
    if args.prefix is not None:
        prefix = args.prefix
    os.makedirs(outdir, exist_ok=True)

    info, body = parse_eml(eml_path)
    account, summary, txns = extract_summary_and_txns(TableExtractor.from_html(body).tables)
    if not summary:
        raise ValueError('未能解析到账务说明部分，无法核对。')

    # 1) 核对：交易明细必须与账务说明一致
    try:
        sums, counts = verify(summary, txns)
    except ValueError as e:
        print(f'错误: 核对失败: {e}', file=sys.stderr)
        return 1

    # 2) 按卡汇总，并验证分卡合计 == 全账户合计
    cards = group_by_card(txns)
    due_date = account.get('到期还款日', '')
    card_rows, card_summary, totals = build_card_summary(cards, due_date)
    verify_card_totals(sums, totals)

    # 3) 输出 xlsx
    paths, card_paths = write_outputs(outdir, prefix, info, summary, txns,
                                      cards, card_rows, card_summary)

    # 4) 报告
    print_report(paths, card_paths, txns, counts, sums, card_rows, totals, summary)
    return 0


if __name__ == '__main__':
    sys.exit(main())
